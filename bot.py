from telegram import (
    Update,
    InlineKeyboardButton,
    InlineKeyboardMarkup
)
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    CallbackQueryHandler,
    MessageHandler,
    filters,
    ContextTypes
)

from structure import MENU_STRUCTURE
from content import CONTENT

from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

import os
TOKEN = os.getenv("BOT_TOKEN")

EXCEL_FILE = "suggestions.xlsx"


# -------------------- СОХРАНЕНИЕ В EXCEL --------------------

def save_to_excel(user, text):
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Предложения"
        ws.append(["Дата", "Username", "User ID", "Текст"])
        wb.save(EXCEL_FILE)

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        user.username,
        user.id,
        text
    ])

    wb.save(EXCEL_FILE)


# -------------------- ГЛАВНОЕ МЕНЮ --------------------

async def show_start(update: Update, context: ContextTypes.DEFAULT_TYPE, edit=False):
    keyboard = [
        [InlineKeyboardButton("🎓 Студент", callback_data="user_student")],
        [InlineKeyboardButton("🏗 Практикующий инженер", callback_data="user_engineer")],
        [InlineKeyboardButton("📐 Инженер старой школы", callback_data="user_oldschool")],
        [InlineKeyboardButton("💬 Предложения", callback_data="suggestions")]
    ]

    text = (
        "Добро пожаловать в StructAI.\n"
        "Это учебный и справочный бот по Еврокодам (СП РК EN).\n\n"
        "Здесь вы можете быстро найти разделы нормативов, формулы, "
        "комбинации нагрузок и основные положения расчёта.\n\n"
        "В дальнейшем планируется внедрение интеллектуального помощника, "
        "который поможет ориентироваться в Еврокодах, находить нужные пункты, "
        "разъяснять требования и подсказывать по вопросам расчёта и проектирования.\n\n"
        "Цель бота — упростить изучение Еврокодов и сделать работу с ними "
        "более удобной и понятной.\n\n"
        "Пожалуйста, ответьте, кто Вы?"
    )

    if edit:
        await update.callback_query.edit_message_text(
            text,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
    else:
        await update.message.reply_text(
            text,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await show_start(update, context)


# -------------------- CALLBACK --------------------

async def handle_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data

    # ---------------- ПРЕДЛОЖЕНИЯ ----------------

    if data == "suggestions":
        context.user_data["suggest_mode"] = True
        keyboard = [
            [InlineKeyboardButton("⬅ Назад", callback_data="back_start")],
            [InlineKeyboardButton("🏠 В главное меню", callback_data="back_start")]
        ]
        await query.edit_message_text(
            "Напишите ваше предложение по улучшению StructAI:",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        return

    # ---------------- РОЛЬ ----------------

    if data.startswith("user_"):
        keyboard = [
            [InlineKeyboardButton("📘 Изучать нормы поэтапно", callback_data="mode_study")],
            [InlineKeyboardButton("🤖 Задать вопрос по Еврокодам", callback_data="mode_question")],
            [InlineKeyboardButton("⬅ Назад", callback_data="back_start")],
            [InlineKeyboardButton("🏠 В главное меню", callback_data="back_start")]
        ]
        await query.edit_message_text(
            "Что Вы хотите?",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

    # ---------------- УЧЕБНЫЙ МОДУЛЬ ----------------

    elif data == "mode_study":
        keyboard = [
            [InlineKeyboardButton("🧩 Структура Еврокодов", callback_data="eu_structure")],
            [InlineKeyboardButton("📚 Выбрать Еврокод", callback_data="choose_eurocode")],
            [InlineKeyboardButton("⬅ Назад", callback_data="user_student")],
            [InlineKeyboardButton("🏠 В главное меню", callback_data="back_start")]
        ]
        await query.edit_message_text(
            "Учебный модуль",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

    elif data == "eu_structure":
        text = CONTENT.get("EU_STRUCTURE", "Текст пока не добавлен.")
        keyboard = [
            [InlineKeyboardButton("⬅ Назад", callback_data="mode_study")],
            [InlineKeyboardButton("🏠 В главное меню", callback_data="back_start")]
        ]
        await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard))

    elif data == "mode_question":
    context.user_data["ai_mode"] = True
    keyboard = [
        [InlineKeyboardButton("⬅ Назад", callback_data="user_student")],
        [InlineKeyboardButton("🏠 В главное меню", callback_data="back_start")]
    ]
        await query.edit_message_text(
        "Напишите ваш вопрос по Еврокодам:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )


    # ---------------- ВЫБОР ЕВРОКОДА ----------------

    elif data == "choose_eurocode":
        keyboard = [
            [InlineKeyboardButton("EN 1990 — Основы проектирования", callback_data="en1990_main")],
            [InlineKeyboardButton("⬅ Назад", callback_data="mode_study")],
            [InlineKeyboardButton("🏠 В главное меню", callback_data="back_start")]
        ]
        await query.edit_message_text("Выбери Еврокод", reply_markup=InlineKeyboardMarkup(keyboard))

    # ---------------- EN1990 ----------------

    elif data == "en1990_main":
        keyboard = [
            [InlineKeyboardButton("❓ Что такое EN 1990", callback_data="content_EN1990_about|en1990_main")],
            [InlineKeyboardButton("🎯 Зачем он нужен", callback_data="content_EN1990_purpose|en1990_main")],
            [InlineKeyboardButton("📑 Структура EN 1990", callback_data="content_EN1990_structure|en1990_main")],
            [InlineKeyboardButton("▶ Начать изучение", callback_data="en1990_sections")],
            [InlineKeyboardButton("⬅ Назад", callback_data="choose_eurocode")],
            [InlineKeyboardButton("🏠 В главное меню", callback_data="back_start")]
        ]
        await query.edit_message_text(
            "EN 1990 — Основы проектирования",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

    # ---------------- РАЗДЕЛЫ ----------------

    elif data == "en1990_sections":
        sections = MENU_STRUCTURE["EN1990"]["sections"]
        keyboard = []

        for sec_id, sec in sections.items():
            keyboard.append([
                InlineKeyboardButton(sec["title"], callback_data=f"section_{sec_id}")
            ])

        keyboard.append([InlineKeyboardButton("⬅ Назад", callback_data="en1990_main")])
        keyboard.append([InlineKeyboardButton("🏠 В главное меню", callback_data="back_start")])

        await query.edit_message_text(
            "Разделы EN 1990",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

    # ---------------- ПОДРАЗДЕЛЫ ----------------

    elif data.startswith("section_"):
        sec_id = data.replace("section_", "")
        section = MENU_STRUCTURE["EN1990"]["sections"].get(sec_id)

        keyboard = []

        for sub_key, sub_title in section["subsections"].items():
            keyboard.append([
                InlineKeyboardButton(
                    sub_title,
                    callback_data=f"content_{sub_key}|section_{sec_id}"
                )
            ])

        keyboard.append([InlineKeyboardButton("⬅ Назад", callback_data="en1990_sections")])
        keyboard.append([InlineKeyboardButton("🏠 В главное меню", callback_data="back_start")])

        await query.edit_message_text(
            section["title"],
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

    # ---------------- КОНТЕНТ ----------------

    elif data.startswith("content_"):
        payload = data.replace("content_", "")
        key, back_callback = payload.split("|")

        text = CONTENT.get(key, "Текст пока не добавлен.")

        keyboard = [
            [InlineKeyboardButton("⬅ Назад", callback_data=back_callback)],
            [InlineKeyboardButton("🏠 В главное меню", callback_data="back_start")]
        ]

        await query.edit_message_text(
            text,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

    # ---------------- ГЛАВНОЕ МЕНЮ ----------------

    elif data == "back_start":
        context.user_data.clear()
        await show_start(update, context, edit=True)


# -------------------- ОБРАБОТКА ТЕКСТА --------------------

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if context.user_data.get("suggest_mode"):
        save_to_excel(update.message.from_user, update.message.text)
        context.user_data["suggest_mode"] = False
        await update.message.reply_text("Спасибо! Предложение будет учтено ✅")


# -------------------- MAIN --------------------

def main():
    app = ApplicationBuilder().token(TOKEN).build()

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CallbackQueryHandler(handle_callback))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))

    print("StructAI запущен")
    app.run_polling()


if __name__ == "__main__":
    main()

# ============================================================
# ======================= AI МОДУЛЬ ==========================
# ============================================================

from openai import OpenAI
import sqlite3
import numpy as np

OPENAI_KEY = os.getenv("OPENAI_API_KEY")

ai_client = OpenAI(
    api_key=OPENAI_KEY,
    base_url="https://openrouter.ai/api/v1"
)

DB_FILE = "structai_ai.db"

# -------------------- ИНИЦИАЛИЗАЦИЯ БД --------------------

def init_ai_db():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    c.execute("""
        CREATE TABLE IF NOT EXISTS users (
            user_id INTEGER PRIMARY KEY,
            role TEXT
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            question TEXT,
            answer TEXT,
            date TEXT
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS documents (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT,
            content TEXT
        )
    """)

    conn.commit()
    conn.close()

# -------------------- СОХРАНЕНИЕ РОЛИ --------------------

def save_user_role(user_id, role):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("INSERT OR REPLACE INTO users (user_id, role) VALUES (?, ?)", (user_id, role))
    conn.commit()
    conn.close()

def get_user_role(user_id):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("SELECT role FROM users WHERE user_id = ?", (user_id,))
    row = c.fetchone()
    conn.close()
    return row[0] if row else "engineer"

# -------------------- ДОБАВЛЕНИЕ ДОКУМЕНТОВ --------------------

def add_document(title, content):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("INSERT INTO documents (title, content) VALUES (?, ?)", (title, content))
    conn.commit()
    conn.close()

def search_documents(query):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("SELECT content FROM documents WHERE content LIKE ?", (f"%{query}%",))
    results = c.fetchall()
    conn.close()
    return "\n\n".join([r[0][:2000] for r in results[:3]])

# -------------------- СИСТЕМНЫЙ ПРОМПТ --------------------

def build_system_prompt(role):

    base = """
Ты инженерный ассистент по Еврокодам EN 1990–1999,
СП РК EN, НТП РК и национальным приложениям.

Запрещено:
- философия
- медицина
- психология
- темы вне проектирования
- выдуманные нормы

Если вопрос вне нормативов:
ответь: "Вопрос вне области нормативного проектирования."
"""

    if role == "student":
        style = "\nОбъясняй максимально просто, пошагово, с примерами."
    elif role == "engineer":
        style = "\nОтвечай профессионально и технически."
    elif role == "oldschool":
        style = "\nОтвечай технически и при возможности указывай различия со старыми СП."
    else:
        style = ""

    return base + style

# -------------------- ЗАПРОС К ИИ --------------------

async def ask_ai(user_id, question):

    role = get_user_role(user_id)
    system_prompt = build_system_prompt(role)

    docs_context = search_documents(question)

    full_prompt = f"""
Контекст нормативов:
{docs_context}

Вопрос:
{question}
"""

    response = ai_client.chat.completions.create(
        model="mistralai/mistral-7b-instruct",
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": full_prompt}
        ],
        temperature=0.2,
        max_tokens=900
    )

    answer = response.choices[0].message.content

    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute(
        "INSERT INTO history (user_id, question, answer, date) VALUES (?, ?, ?, ?)",
        (user_id, question, answer, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    )
    conn.commit()
    conn.close()

    return answer

# -------------------- ПЕРЕХВАТ CALLBACK ДЛЯ РОЛИ --------------------

old_handle_callback = handle_callback

async def handle_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):

    query = update.callback_query
    data = query.data

    if data == "user_student":
        save_user_role(query.from_user.id, "student")

    elif data == "user_engineer":
        save_user_role(query.from_user.id, "engineer")

    elif data == "user_oldschool":
        save_user_role(query.from_user.id, "oldschool")

    await old_handle_callback(update, context)

# -------------------- РАСШИРЕНИЕ handle_message --------------------

old_handle_message = handle_message

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):

    if context.user_data.get("ai_mode"):
        await update.message.reply_text("Анализ нормативной базы...")
        answer = await ask_ai(update.message.from_user.id, update.message.text)
        await update.message.reply_text(answer)
        return

    await old_handle_message(update, context)

# -------------------- ПЕРЕИНИЦИАЛИЗАЦИЯ MAIN --------------------

old_main = main

def main():
    init_ai_db()
    old_main()
